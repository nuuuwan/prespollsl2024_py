import json
import os
from functools import cache, cached_property

from gig import Ent, EntType, GIGTable
from openpyxl import load_workbook
from utils import File, JSONFile, Log

log = Log('ExcelResultsXlsx')


def parse_int(x):
    try:
        return int(float(x))
    except ValueError:
        log.error(f'Could not parse int: {x}')
        return 0


class ElectionResultsXlsxBase:
    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path

    @cached_property
    def data_list(self) -> list[dict]:
        wb = load_workbook(self.xlsx_path)
        ws = wb['results']
        data_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {
                'pd_id': row[0],
                'ed_name': row[1],
                'pd_name': row[2],
                # space
                'result_time': row[4],
                # space
                'summary': {
                    'electors': row[6],
                    'polled': row[7],
                    'rejected': row[8],
                    'valid': row[9],
                },
                # space
                'subset_party_to_votes': {
                    party_id: row[11 + i]
                    for i, party_id in enumerate(self.PARTY_IDS)
                },
            }

            data_list.append(data)

        return data_list

    @cache
    def get_electors_previous(self, pd_id) -> dict:
        gig_table = GIGTable(
            'government-elections-parliamentary', 'regions-ec', '2020'
        )
        idx = gig_table.remote_data_idx
        return parse_int(idx[pd_id]['electors'])

    def validate_data(self, data: dict) -> list[str]:
        errors = []
        pd_id = data['pd_id']
        data['ed_name']
        data['pd_name']
        data['result_time']
        summary = data['summary']
        subset_party_to_votes = data['subset_party_to_votes']

        if summary['electors'] < summary['polled']:
            errors.append(f'{pd_id}: electors < polled')

        if summary['polled'] < summary['rejected']:
            errors.append(f'{pd_id}: polled < rejected')

        if summary['polled'] < summary['valid']:
            errors.append(f'{pd_id}: polled < valid')

        if summary['valid'] + summary['rejected'] != summary['polled']:
            errors.append(f'{pd_id}: valid + rejected != polled')

        # subset_party_to_votes
        subset_party_total_votes = sum(subset_party_to_votes.values())
        p_subset_party = subset_party_total_votes / summary['valid']

        if subset_party_total_votes > summary['valid']:
            errors.append(f'{pd_id}: subset party votes > valid')

        MIN_ABS_P_SUBSET_PARTY = 0.8
        if p_subset_party < MIN_ABS_P_SUBSET_PARTY:
            errors.append(
                f'{pd_id}: subset party votes out of range ({subset_party_total_votes:,} / {summary["valid"]:,})'
            )

        # previous data
        electors_previous = self.get_electors_previous(pd_id)

        p_elector_growth = (
            summary['electors'] - electors_previous
        ) / electors_previous
        MAX_ABS_P_ELECTOR_GROWTH = 0.1
        if abs(p_elector_growth) > MAX_ABS_P_ELECTOR_GROWTH:
            errors.append(
                f'{pd_id}: elector growth out of range ({electors_previous:,} -> {summary["electors"]:,})'
            )

        return errors

    def validate(self):
        data_list = self.data_list
        assert len(data_list) == 160 + 22

        data_idx = {data['pd_id']: data for data in data_list}
        pds = Ent.list_from_type(EntType.PD)
        eds = Ent.list_from_type(EntType.ED)
        errors = []

        for pd in pds:
            if pd.id not in data_idx:
                errors.append(f'PD {pd.id}: not found')

            if pd.name != data_idx[pd.id]['pd_name']:
                errors.append(f'PD {pd.id}: name mismatch')

        for ed in eds:
            postal_pd_id = ed.id + 'P'
            if postal_pd_id not in data_idx:
                errors.append(f'Postal PD {postal_pd_id} not found')
            if f'Postal - {ed.name}' != data_idx[postal_pd_id]['pd_name']:
                errors.append(f'Postal PD {postal_pd_id}: name mismatch')

        for data in data_list:
            errors.extend(self.validate_data(data))

        if errors:
            log.warning(f'🛑 Found {len(errors)} errors')
            for error in errors:
                log.error("\t" + error)
        else:
            log.info('✅ No errors found.')

        return errors

    def write_json(
        self,
    ):
        path_prefix = self.xlsx_path[:-5].upper()
        json_path = path_prefix + '.json'
        JSONFile(json_path).write(self.data_list)
        n = len(self.data_list)
        log.info(f'Wrote {n} items to {json_path}')
        return json_path

    def write_js(self):
        path_prefix = self.xlsx_path[:-5].upper()
        js_path = path_prefix + '.js'
        var_name = os.path.split(path_prefix)[-1]

        var_value_str = json.dumps(self.data_list, indent=2)

        lines = [
            '// Auto-generated by ElectionResultsXlsx',
            '',
            f'const {var_name} = {var_value_str};',
            '',
            f'export default {var_name};',
            '',
        ]
        File(js_path).write_lines(lines)
        log.info(f'Wrote {len(self.data_list)} items to {js_path}')
        return js_path
