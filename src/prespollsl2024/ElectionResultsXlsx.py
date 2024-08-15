import os
import random

from gig import Ent, EntType, GIGTable
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from utils import Log, Time, TimeFormat

log = Log('ExcelResultsXlsx')


def random_float(min_value, max_value):
    return random.random() * (max_value - min_value) + min_value


class ElectionResultsXlsx:
    GIG_TABLE_ELECTION_2020 = GIGTable(
        'government-elections-parliamentary', 'regions-ec', '2020'
    )

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path

    @staticmethod
    def add_header(ws, party_ids):
        fields = (
            [
                'pd_id',
                'ed_name',
                'pd_name',
                'electors_2020',
                "",
                'result_time',
                'electors',
                'polled',
                'rejected',
                'valid',
            ]
            + party_ids
            + [
                '',
                # '_OTHERS',
                # '_electors',
                # '_polled',
                # '_rejected',
                # '_valid',
                # '_sum',
            ]
        )
        ws.append(fields)

    @staticmethod
    def add_statistics(row, party_ids, electors_previous):
        electors = int(electors_previous * random_float(1, 1.1))
        polled = int(electors * random_float(0.6, 0.9))
        rejected = int(polled * random_float(0.01, 0.03))
        valid = polled - rejected
        result_time = TimeFormat('%Y-%m-%d %H:%M').format(
            Time(Time.now().ut - random.random() * 3600 * 12)
        )

        row.extend(["", result_time, electors, polled, rejected, valid])

        party_weight = [random_float(0.1, 0.5) for _ in party_ids]
        total_weight = sum(party_weight)

        for i, _ in enumerate(party_ids):
            p_party = 0.95 * party_weight[i] / total_weight
            row.append(int(valid * p_party))

    @staticmethod
    def add_auto_columns(i, row):
        row.append("")
        # row.append(f'=J{i} - SUM(K{i}:N{i})')
        # row.append(f'=IF(G{i} < D{i}, "DOWN", "")')
        # row.append(f'=IF(H{i} > G{i}, "UP", "")')
        # row.append(f'=IF(I{i} > H{i}, "UP", "")')
        # row.append(f'=IF(J{i} < I{i}, "DOWN", "")')
        # row.append(f'=IF(J{i} + I{i} <> H{i}, "UNEQUAL", "")')

    @staticmethod
    def add_eds(ws, party_ids):
        eds = Ent.list_from_type(EntType.ED)
        election_2020_idx = (
            ElectionResultsXlsx.GIG_TABLE_ELECTION_2020.remote_data_idx
        )
        for i_row, ed in enumerate(eds, start=2):
            postal_pd_id = ed.id + 'P'
            data = election_2020_idx[postal_pd_id]
            electors_2020 = int(round(float(data['electors']), 0))

            row = [
                postal_pd_id[3:],
                ed.name,
                'Postal - ' + ed.name,
                electors_2020,
            ]
            ElectionResultsXlsx.add_statistics(row, party_ids, electors_2020)
            ElectionResultsXlsx.add_auto_columns(i_row, row)
            ws.append(row)

    @staticmethod
    def add_pds(ws, party_ids):
        pds = Ent.list_from_type(EntType.PD)
        for i_row, pd in enumerate(pds, start=2):
            ed_id = pd.id[:-1]
            ed = Ent.from_id(ed_id)
            electors_2020 = int(
                round(
                    pd.gig(
                        ElectionResultsXlsx.GIG_TABLE_ELECTION_2020
                    ).electors,
                    0,
                )
            )
            row = [
                pd.id[3:],
                ed.name,
                pd.name,
                electors_2020,
            ]
            ElectionResultsXlsx.add_statistics(row, party_ids, electors_2020)
            ElectionResultsXlsx.add_auto_columns(i_row, row)
            ws.append(row)

    @staticmethod
    def fix_column_widths(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_value = str(cell.value)
                    if cell_value.startswith('='):
                        continue
                    if len(cell_value) > max_length:
                        max_length = len(str(cell.value))
                except BaseException:
                    pass
            adjusted_width = max_length
            ws.column_dimensions[column_letter].width = max(
                10, adjusted_width
            )

            ws.column_dimensions["F"].width = 16
            for column_letter in ['E', "O"]:
                ws.column_dimensions[column_letter].width = 4

    @staticmethod
    def format_columns(ws):
        number_format = '#,##0'
        for col in ws.iter_cols(min_col=4):
            for cell in col:
                cell.number_format = number_format

        for col in ws.iter_cols(min_col=6, max_col=6):
            for cell in col:
                cell.number_format = 'yyyy-mm-dd hh-mm'

    @staticmethod
    def format_cells(
        ws,
    ):
        fill = PatternFill(
            start_color="000088", end_color="000088", fill_type="solid"
        )
        font = Font(color="FFFFFF")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font

    @staticmethod
    def freeze(ws):
        ws.freeze_panes = 'A2'

    @staticmethod
    def build(date, party_ids):
        xlsx_path = os.path.join('data', f'election-{date}.xlsx')
        log.debug(f'Building {xlsx_path}...')
        wb = Workbook()
        ws = wb.active
        ws.title = 'results'

        ElectionResultsXlsx.add_header(ws, party_ids)
        ElectionResultsXlsx.add_eds(ws, party_ids)
        ElectionResultsXlsx.add_pds(ws, party_ids)

        ElectionResultsXlsx.format_columns(ws)
        ElectionResultsXlsx.fix_column_widths(ws)
        ElectionResultsXlsx.format_cells(ws)
        ElectionResultsXlsx.freeze(ws)

        wb.save(xlsx_path)
        log.info(f'Wrote {xlsx_path}')


        return ElectionResultsXlsx(xlsx_path)

    @staticmethod
    def load(date, party_ids):
        xlsx_path = os.path.join('data', f'election-{date}.xlsx')
        if os.path.exists(xlsx_path):
            log.warning(f'{xlsx_path} exists')
            return ElectionResultsXlsx(xlsx_path)
        
        return ElectionResultsXlsx.build(date, party_ids)


if __name__ == '__main__':
    ElectionResultsXlsx.load(
        '2024-09-21',
        [
            'SJB',
            'NPP',
            'UNP',
            'SLPP',
        ],
    )
