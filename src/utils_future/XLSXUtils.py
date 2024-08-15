import random

from openpyxl.styles import Font, PatternFill
from utils import Log

log = Log('ExcelResultsXlsx')


def random_float(min_value, max_value):
    return random.random() * (max_value - min_value) + min_value


class XLSXUtils:
    @staticmethod
    def fix_column_widths(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_value = str(cell.value)
                    if cell_value.startswith('='):
                        continue
                    if len(cell_value) > max_length:
                        max_length = len(str(cell.value))
                except BaseException:
                    pass
            adjusted_width = max_length
            ws.column_dimensions[column_letter].width = max(8, adjusted_width)

    @staticmethod
    def format_cells(
        ws,
    ):
        fill = PatternFill(
            start_color="000088", end_color="000088", fill_type="solid"
        )
        font = Font(color="FFFFFF")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font

    @staticmethod
    def freeze(ws):
        ws.freeze_panes = 'A2'
